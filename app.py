# app.py (v2 — Lightweight, HF Spaces ready)

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import io
import os
import re
import uuid
import threading
import pandas as pd
import json
from scraper import fetch_vtu_results 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import db

import sys
import builtins

def force_print(*args, **kwargs):
    kwargs['file'] = sys.stderr
    kwargs['flush'] = True
    builtins.print(*args, **kwargs)

print = force_print
# Initialize Database
db.init_db()

JOBS = {}
app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app, resources={r"/*": {"origins": "*"}})
CREDIT_MAP = {
    'BCS401': 3, # ADA
    'BAD402': 4, # AI
    'BCS403': 4, # DBMS
    'BCS405A': 3, #DMS
    'BCS405C': 3, #OT
    
    
    'BCSL404': 1, #ADAL
    'BDSL456B': 1, #MONGO
    'BDSL456C': 1, #MERN
    
    
    'BBOC407': 2, #BIO
    'BUHK408': 1, #UHV
    'BPEK459': 0, #PE
    'BYOK459': 0, #YOGA
    'BNSK459': 0, #NSS
    'BAI601':4,
    'BAI602':4,
    'BAI685':2,
    'BAI613A':3,
    'BAI613D':3,
    'BXX654X':3,
    'BAIL606':1,
    'BAIL657C':2,
    'BAI657D':2,
    'BIKS609':0,
    'BNSK658':0,
    'BPEK658':0,
    'BYOK658':0
}

# Seed if empty, then load dynamic map from DB
db.seed_credits_if_empty(CREDIT_MAP)
CREDIT_MAP = db.get_all_credits()

def get_grade_point(marks_str, result_str):
    if result_str in ['F', 'A', 'NE']: return 0
    try:
        marks = int(marks_str)
        if 90 <= marks <= 100: return 10
        elif 80 <= marks <= 89: return 9
        elif 70 <= marks <= 79: return 8
        elif 60 <= marks <= 69: return 7
        elif 55 <= marks <= 59: return 6
        elif 50 <= marks <= 54: return 5
        elif 40 <= marks <= 49: return 4
        else: return 0
    except (ValueError, TypeError): return 0
def generate_usn_range(start_usn, end_usn):
    try:
        start_prefix, end_prefix = start_usn[:-3], end_usn[:-3]
        if start_prefix != end_prefix: return []
        start_num, end_num = int(start_usn[-3:]), int(end_usn[-3:])
        return [f"{start_prefix}{str(i).zfill(3)}" for i in range(start_num, end_num + 1)]
    except (ValueError, IndexError): return []
def format_data_for_wide_export(results_data):
    if not results_data: return [], []
    all_subject_codes = set(sub['subject_code'] for student in results_data for sub in student.get('subjects', []))
    elective_groups = {}
    for code in all_subject_codes:
        if not code: continue
        
        # Adaptive grouping for B-scheme electives (e.g., BCS603A, BME654A, BYOK658, BNSS658)
        match_bxx = re.match(r'^B[A-Z0-9]{2,3}(\d{3})[A-Z]?$', code)
        # Grouping for 18/21 scheme electives (e.g., 18CS641)
        match_old = re.match(r'^(\d{2}[A-Z]{2,3}\d{2})\d$', code)
        
        if match_bxx:
            group_key = f"BXX{match_bxx.group(1)}"
        elif match_old:
            group_key = match_old.group(1)
        elif code and code[-1].isalpha():
            group_key = code[:-1]
        else:
            group_key = code
            
        if group_key not in elective_groups: elective_groups[group_key] = []
        elective_groups[group_key].append(code)
    display_headers = []
    for key, codes in sorted(elective_groups.items()):
        is_elective = len(codes) > 1
        header_info = {"header": key if is_elective else codes[0], "is_elective": is_elective}
        display_headers.append(header_info)

    processed_records = []
    for student in results_data:
        student_subjects = {s['subject_code']: s for s in student.get('subjects', [])}
        record = {'USN': student.get('usn', 'N/A'),'Name': student.get('student_name', 'N/A'),'subjects_data': {}}
        for header_info in display_headers:
            header_key = header_info["header"]
            found_subject = None
            if header_info["is_elective"]:
                for elective_code in elective_groups[header_key]:
                    if elective_code in student_subjects: found_subject = student_subjects[elective_code]; break
            else:
                if header_key in student_subjects: found_subject = student_subjects[header_key]
            if found_subject:
                record['subjects_data'][header_key] = {'Course': found_subject.get('subject_code', '-'), 'IA': found_subject.get('internal_marks', '-'),'Ex': found_subject.get('external_marks', '-'), 'Total': found_subject.get('total', '-'), 'Pass/Fail': found_subject.get('result', '-')}
            else:
                record['subjects_data'][header_key] = {'Course': '-', 'IA': '-', 'Ex': '-', 'Total': '-', 'Pass/Fail': '-'}
        
        # Stats are already calculated in the raw data, just copy them over
        record['sgpa'] = student.get('sgpa', 'N/A')
        record['percentage'] = student.get('percentage', 'N/A')
        record['class'] = student.get('class', 'N/A')
        record['subjects_failed'] = student.get('subjects_failed', 0)
        record['subjects_absent'] = student.get('subjects_absent', 0)
        
        processed_records.append(record)
    return processed_records, display_headers

def calculate_student_stats(student, credit_map=None):
    """Calculates SGPA, percentage, and pass/fail status for a single student dictionary."""
    if credit_map is None:
        credit_map = db.get_all_credits() # Fetch from DB directly if not provided

    total_credit_points, total_grade_credit_product = 0, 0
    num_subjects = len(student.get('subjects', []))
    max_possible_marks = num_subjects * 100
    total_marks_obtained = 0
    has_failed_a_subject = False
    subjects_failed = 0
    subjects_absent = 0

    print(f"\n[DEBUG SGPA] === Calculating stats for {student.get('usn', 'Unknown')} ===")
    print(f"[DEBUG SGPA] Subjects list length: {num_subjects}")

    for subject in student.get('subjects', []):
        res = subject.get('result', '')
        if res == 'F': 
            has_failed_a_subject = True
            subjects_failed += 1
        elif res == 'A':
            subjects_absent += 1
            
        try: total_marks_obtained += int(subject.get('internal_marks', '0'))
        except: pass
        try: total_marks_obtained += int(subject.get('external_marks', '0'))
        except: pass

        subj_code = subject.get('subject_code')
        if not subj_code: continue
        
        # --- Smart Subject Resolution & Auto-Discovery ---
        credits = credit_map.get(subj_code)
        
        # If not found directly, check if user added a wildcard group (BXXX or BXX)
        if credits is None:
            match_bxx = re.match(r'^B[A-Z0-9]{2,3}(\d{3})[A-Z]?$', subj_code)
            if match_bxx:
                num_code = match_bxx.group(1)
                group_key_3x = f"BXXX{num_code}" # e.g., BXXX658
                group_key_2x = f"BXX{num_code}"  # e.g., BXX658
                
                if group_key_3x in credit_map:
                    credits = credit_map[group_key_3x]
                elif group_key_2x in credit_map:
                    credits = credit_map[group_key_2x]
        
        # Temporarily cache this mapping so we don't repeat the regex
        credit_map[subj_code] = credits
        
        # Override with scraped credits if they were found in the HTML table
        scraped_credits = subject.get('credits')
        if scraped_credits:
            try:
                credits = int(scraped_credits)
                print(f"[DEBUG SGPA] {subj_code}: Scraped credits found = {credits}")
            except ValueError:
                pass

        # If still not found anywhere, ask the user
        if credits is None:
            print(f"[MISSING CREDITS] Unknown subject '{subj_code}' encountered.")
            if 'missing_subjects' not in student:
                student['missing_subjects'] = []
            if subj_code not in student['missing_subjects']:
                student['missing_subjects'].append(subj_code)
            credits = 0
            
        # Temporarily cache this mapping so we don't repeat the regex
        credit_map[subj_code] = credits

        if credits is not None:
            gp_scraped = subject.get('grade_point')
            if gp_scraped:
                try:
                    grade_point = float(gp_scraped)
                    print(f"[DEBUG SGPA] {subj_code}: Scraped Grade Point = {grade_point}")
                except ValueError:
                    grade_point = get_grade_point(subject.get('total'), res)
            else:
                grade_point = get_grade_point(subject.get('total'), res)
                
            print(f"[DEBUG SGPA] {subj_code}: Credits = {credits}, Grade Point = {grade_point}, Result = {res}, Total Marks = {subject.get('total')}")

            total_credit_points += credits
            total_grade_credit_product += (grade_point * credits)

    sgpa_str = f"{(total_grade_credit_product / total_credit_points):.2f}" if total_credit_points > 0 else "N/A"
    print(f"[DEBUG SGPA] === FINAL SGPA: {sgpa_str} (Total Credit Points: {total_credit_points}, Grade*Credit: {total_grade_credit_product}) ===\n")
    
    student['sgpa'] = sgpa_str
    percentage = (total_marks_obtained / max_possible_marks) * 100 if max_possible_marks > 0 else 0
    student['percentage'] = f"{percentage:.2f}%" if max_possible_marks > 0 else "N/A"
    
    if has_failed_a_subject or percentage < 50: student['class'] = "FAIL"
    elif percentage >= 70: student['class'] = "FCD"
    elif percentage >= 60: student['class'] = "FC"
    elif percentage >= 50: student['class'] = "SC"
    else: student['class'] = "N/A" if student['percentage'] == "N/A" else "FAIL"
    
    student['subjects_failed'] = subjects_failed
    student['subjects_absent'] = subjects_absent
    student['total_marks'] = total_marks_obtained
    return student
# ... (all routes up to /analyze are the same) ...
@app.route('/')
def index():
    return jsonify({"status": "Vorniity API is running! Go to vorniity.com to use the dashboard."})
@app.route('/api/scrape_chunk', methods=['POST'])
def scrape_chunk():
    """
    Synchronous endpoint to scrape a chunk of USNs.
    Designed to be called by the frontend orchestrator.
    """
    data = request.json
    usn_list = data.get('usns', [])
    vtu_url = data.get('vtu_url', '')
    
    if not usn_list or not vtu_url:
        return jsonify({'error': 'Missing usns or vtu_url'}), 400
        
    try:
        # Load fresh credits for accurate SGPA
        fresh_credits = db.get_all_credits()
        if not fresh_credits:
            fresh_credits = CREDIT_MAP.copy() # Fallback if DB limit is hit
            
        results, skipped = fetch_vtu_results(usn_list, vtu_url, job_state=None)
        
        missing_subjects = set()
        if results:
            results = [calculate_student_stats(r, fresh_credits) for r in results]
            for r in results:
                if 'missing_subjects' in r:
                    missing_subjects.update(r['missing_subjects'])
            
        return jsonify({
            'success': True, 
            'results': results, 
            'skipped': skipped,
            'missing_subjects': list(missing_subjects)
        })
    except Exception as e:
        print(f"[SCRAPE CHUNK ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/students/all', methods=['GET'])
def get_all_students_from_db():
    """Fetches the latest scraped profile for every unique student in the database."""
    try:
        conn = db.get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT usn, data, MAX(timestamp) as max_time FROM results_cache_v2 GROUP BY usn, data ORDER BY usn")
            rows = cursor.fetchall()
        conn.close()
        
        fresh_credits = db.get_all_credits()
        students = []
        for row in rows:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            data = calculate_student_stats(data, fresh_credits)
            students.append(data)
            
        return jsonify(students)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/credits', methods=['GET', 'POST'])
def manage_credits():
    """API endpoint to get or update subject credits."""
    global CREDIT_MAP
    if request.method == 'GET':
        return jsonify(db.get_all_credits())
    
    if request.method == 'POST':
        data = request.json
        code = data.get('subject_code')
        credits = data.get('credits')
        
        if not code or credits is None:
            return jsonify({'error': 'Missing subject_code or credits'}), 400
            
        try:
            credits = int(credits)
        except ValueError:
            return jsonify({'error': 'Credits must be an integer'}), 400
            
        success = db.save_credit(code.upper(), credits)
        if success:
            # Update the global in-memory map so we don't need to restart
            CREDIT_MAP[code.upper()] = credits
            return jsonify({'success': True, 'subject_code': code.upper(), 'credits': credits})
        else:
            return jsonify({'error': 'Database error'}), 500

@app.route('/api/credits/<subject_code>', methods=['DELETE'])
def delete_credit_route(subject_code):
    success = db.delete_credit(subject_code)
    if success:
        CREDIT_MAP.pop(subject_code.upper(), None)
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to delete credit'}), 500

@app.route('/api/classes', methods=['GET', 'POST'])
def manage_classes():
    if request.method == 'GET':
        return jsonify(db.get_all_classes())
        
    if request.method == 'POST':
        data = request.json
        name, start_usn, end_usn = data.get('name'), data.get('start_usn'), data.get('end_usn')
        if not all([name, start_usn, end_usn]):
            return jsonify({'error': 'Missing required fields'}), 400
            
        success = db.create_class(name, start_usn, end_usn)
        if success: return jsonify({'success': True})
        else: return jsonify({'error': 'Failed to create class'}), 500

@app.route('/api/classes/<int:class_id>', methods=['DELETE', 'PUT'])
def handle_class_by_id(class_id):
    if request.method == 'DELETE':
        success = db.delete_class(class_id)
        if success: return jsonify({'success': True})
        else: return jsonify({'error': 'Failed to delete class'}), 500
        
    if request.method == 'PUT':
        data = request.json
        name, start_usn, end_usn = data.get('name'), data.get('start_usn'), data.get('end_usn')
        if not all([name, start_usn, end_usn]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        success = db.update_class(class_id, name, start_usn, end_usn)
        if success: return jsonify({'success': True})
        else: return jsonify({'error': 'Failed to update class'}), 500

@app.route('/api/student/<usn>', methods=['DELETE'])
def delete_student_route(usn):
    success = db.delete_student(usn)
    if success: return jsonify({'success': True})
    else: return jsonify({'error': 'Failed to delete student'}), 500

@app.route('/api/class/<int:class_id>/students', methods=['GET'])
def get_class_students(class_id):
    classes = db.get_all_classes()
    target_class = next((c for c in classes if c['id'] == class_id), None)
    if not target_class: return jsonify({'error': 'Class not found'}), 404
    
    usn_list = generate_usn_range(target_class['start_usn'], target_class['end_usn'])
    if not usn_list: return jsonify({'error': 'Invalid USN range in class definition'}), 400
    
    try:
        conn = db.get_db_connection()
        with conn.cursor() as cursor:
            # We need the most recent scrape for each USN in the list
            placeholders = ','.join(['%s'] * len(usn_list))
            query = f"SELECT usn, data, MAX(timestamp) as max_time FROM results_cache_v2 WHERE usn IN ({placeholders}) GROUP BY usn, data ORDER BY usn"
            cursor.execute(query, usn_list)
            rows = cursor.fetchall()
        conn.close()
        
        fresh_credits = db.get_all_credits()
        students = []
        for row in rows:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            data = calculate_student_stats(data, fresh_credits)
            students.append(data)
            
        return jsonify(students)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def get_stats():
    try:
        conn = db.get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(DISTINCT usn) as c FROM results_cache_v2")
            total_students = cursor.fetchone()['c'] or 0
            
            cursor.execute("SELECT COUNT(id) as c FROM classes")
            total_classes = cursor.fetchone()['c'] or 0
        conn.close()
        
        return jsonify({
            'total_students': total_students,
            'total_classes': total_classes
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/<usn>', methods=['GET'])
def get_student_history(usn):
    """Fetches all cached semesters for a specific student."""
    usn = usn.upper()
    try:
        conn = db.get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT url, data, timestamp FROM results_cache_v2 WHERE usn = %s ORDER BY timestamp DESC", (usn,))
            rows = cursor.fetchall()
        conn.close()
        
        fresh_credits = db.get_all_credits()
        history = []
        for row in rows:
            data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            data = calculate_student_stats(data, fresh_credits)
            data['scraped_url'] = row['url']
            data['scraped_at'] = row['timestamp'].isoformat() if hasattr(row['timestamp'], 'isoformat') else str(row['timestamp'])
            history.append(data)
            
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/history/scrapes', methods=['GET'])
def get_scrapes():
    """Fetches all scrape history."""
    try:
        history = db.get_scrape_history()
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/history/save', methods=['POST'])
def save_history():
    """Saves a scrape job to the history database."""
    try:
        data = request.json
        import uuid
        job_id = str(uuid.uuid4())
        success = db.save_scrape_history(
            job_id,
            data.get('start_usn', ''),
            data.get('end_usn', ''),
            data.get('total_usns', 0),
            data.get('completed', 0),
            data.get('time_taken', 0),
            data.get('status', 'Completed')
        )
        if success:
            return jsonify({'success': True}), 200
        return jsonify({'error': 'Database error'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/download/excel', methods=['POST'])
def download_excel():
    results_data = request.json
    records, display_headers = format_data_for_wide_export(results_data)
    if not records: return "No data to export", 400
    wb = Workbook()
    ws = wb.active; ws.title = "Results"
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=1, column=1, value='USN').font = header_font; ws.merge_cells('A1:A2')
    ws.cell(row=1, column=2, value='Name').font = header_font; ws.merge_cells('B1:B2')
    col_idx = 3
    for header_info in display_headers:
        is_elective = header_info["is_elective"]
        colspan = 5 if is_elective else 4
        ws.cell(row=1, column=col_idx, value=header_info["header"]).font = header_font
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + colspan - 1)
        sub_headers = ['Course', 'IA', 'Ex', 'Total', 'Pass/Fail'] if is_elective else ['IA', 'Ex', 'Total', 'Pass/Fail']
        for i, sub_header in enumerate(sub_headers):
            ws.cell(row=2, column=col_idx + i, value=sub_header).font = header_font
        col_idx += colspan
    summary_start_col = col_idx
    summary_headers = ['NO OF SUBJECTS FAILED', 'NO OF SUBJECTS ABSENT', 'Percentage', 'Class', 'SGPA']
    for i, h in enumerate(summary_headers):
        ws.cell(row=1, column=summary_start_col + i, value=h).font = header_font
        ws.merge_cells(start_row=1, start_column=summary_start_col + i, end_row=2, end_column=summary_start_col + i)
    row_idx = 3
    for record in records:
        ws.cell(row=row_idx, column=1, value=record['USN'])
        ws.cell(row=row_idx, column=2, value=record['Name'])
        col_idx = 3
        for header_info in display_headers:
            header_key, is_elective = header_info["header"], header_info["is_elective"]
            data = record['subjects_data'][header_key]
            if is_elective:
                ws.cell(row=row_idx, column=col_idx, value=data['Course']); ws.cell(row=row_idx, column=col_idx + 1, value=data['IA']); ws.cell(row=row_idx, column=col_idx + 2, value=data['Ex']); ws.cell(row=row_idx, column=col_idx + 3, value=data['Total']); pf_cell = ws.cell(row=row_idx, column=col_idx + 4, value=data['Pass/Fail'])
                if data['Pass/Fail'] == 'F': pf_cell.fill = red_fill
                col_idx += 5
            else:
                ws.cell(row=row_idx, column=col_idx, value=data['IA']); ws.cell(row=row_idx, column=col_idx + 1, value=data['Ex']); ws.cell(row=row_idx, column=col_idx + 2, value=data['Total']); pf_cell = ws.cell(row=row_idx, column=col_idx + 3, value=data['Pass/Fail'])
                if data['Pass/Fail'] == 'F': pf_cell.fill = red_fill
                col_idx += 4
        ws.cell(row=row_idx, column=col_idx, value=record['subjects_failed']); ws.cell(row=row_idx, column=col_idx + 1, value=record['subjects_absent']); ws.cell(row=row_idx, column=col_idx + 2, value=record['percentage']); ws.cell(row=row_idx, column=col_idx + 4, value=record['sgpa']); class_cell = ws.cell(row=row_idx, column=col_idx + 3, value=record['class'])
        if record['class'] == 'FAIL': class_cell.fill = red_fill
        row_idx += 1
    for row in ws.iter_rows():
        for cell in row: cell.alignment = center_align; cell.border = thin_border
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='results.xlsx')
def calculate_statistics(df):
    stats = {}
    class_col_name = [col for col in df.columns if 'Class' in col[0]][0]
    class_series = df[class_col_name]
    stats['overall'] = {"fcd": int((class_series == 'FCD').sum()),"fc": int((class_series == 'FC').sum()),"sc": int((class_series == 'SC').sum()),"fail": int((class_series == 'FAIL').sum())}
    subject_stats = {}
    subject_headers = [col[0] for col in df.columns if col[0] not in ['USN', 'Name'] and not col[0].startswith('Unnamed')]
    unique_subject_headers = sorted(list(set(subject_headers)))
    for header in unique_subject_headers:
        if header in subject_stats: continue
        subject_df = df[header]
        pass_fail_series = subject_df['Pass/Fail']
        passes = (pass_fail_series == 'P').sum()
        fails = (pass_fail_series == 'F').sum()
        absent = (pass_fail_series == 'A').sum()
        withheld = (pass_fail_series == 'NE').sum()
        appeared = passes + fails
        pass_percentage = (passes / appeared) * 100 if appeared > 0 else 0
        subject_stats[header] = {"appeared": int(appeared),"pass": int(passes),"fail": int(fails),"absent": int(absent),"withheld": int(withheld),"pass_percentage": pass_percentage}
    stats['subject_wise'] = subject_stats
    return stats
@app.route('/analyze', methods=['GET', 'POST'])
def analyze():
    if request.method == 'POST':
        if 'results_file' not in request.files: return "No file part", 400
        file = request.files['results_file']
        if file.filename == '' or not file.filename.endswith('.xlsx'): return "Please upload a valid .xlsx file", 400
        try:
            df = pd.read_excel(file, header=[0, 1])
            stats = calculate_statistics(df)
            return render_template('analyze.html', stats=stats, stats_json=json.dumps(stats))
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            return "Error processing Excel file. Ensure it is in the correct format.", 500
    return render_template('analyze.html', stats=None)

# --- NEW: Function to Create the Analysis Excel Workbook ---
def create_analysis_workbook(stats):
    wb = Workbook()
    
    # --- Sheet 1: Overall Summary ---
    ws_overall = wb.active
    ws_overall.title = "Overall Summary"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # Add Overall Class Distribution
    ws_overall.append(['Overall Class Distribution'])
    ws_overall['A1'].font = header_font
    headers1 = ['FCD', 'First Class', 'Second Class', 'Fail', 'Total Pass', 'Total Appeared']
    for i, header in enumerate(headers1):
        cell = ws_overall.cell(row=2, column=i+1, value=header)
        cell.font = header_font
        cell.alignment = center_align

    o = stats['overall']
    total_pass = o['fcd'] + o['fc'] + o['sc']
    total_appeared = total_pass + o['fail']
    ws_overall.append([o['fcd'], o['fc'], o['sc'], o['fail'], total_pass, total_appeared])

    # --- Sheet 2: Subject-Wise Analysis ---
    ws_subject = wb.create_sheet("Subject-Wise Analysis")
    
    headers2 = ["Subject", "Appeared", "Pass", "Fail", "Absent", "Withheld / NE", "Pass %"]
    for i, header in enumerate(headers2):
        cell = ws_subject.cell(row=1, column=i+1, value=header)
        cell.font = header_font
        cell.alignment = center_align

    for code, data in stats['subject_wise'].items():
        pass_percent = f"{data['pass_percentage']:.2f}%"
        ws_subject.append([code, data['appeared'], data['pass'], data['fail'], data['absent'], data['withheld'], pass_percent])

    # Auto-fit column widths for both sheets
    for sheet in wb.sheetnames:
        for col in wb[sheet].columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except: pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- NEW: Route to Handle Downloading the Analysis ---
@app.route('/download/analysis', methods=['POST'])
def download_analysis():
    stats_data = request.json
    if not stats_data:
        return "No analysis data provided", 400
    
    try:
        output = create_analysis_workbook(stats_data)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='results_analysis.xlsx'
        )
    except Exception as e:
        print(f"Error creating analysis workbook: {e}")
        return "Error creating analysis file", 500

# --- NEW: Route to Handle Clearing Database ---
@app.route('/api/database/clear', methods=['DELETE'])
def clear_database():
    success, msg = db.clear_database()
    if success:
        return jsonify({'success': True, 'message': msg}), 200
    return jsonify({'success': False, 'message': msg}), 500

@app.route('/api/recalculate', methods=['POST'])
def recalculate_batch():
    students = request.json.get('students', [])
    fresh_credits = db.get_all_credits()
    updated_students = [calculate_student_stats(s, fresh_credits) for s in students]
    return jsonify({'success': True, 'results': updated_students})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 7860))
    app.run(host='0.0.0.0', port=port, debug=True)